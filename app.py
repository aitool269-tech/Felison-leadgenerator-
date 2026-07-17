"""AFM Leadgenerator — import, scoring, verdeling en opvolging van AFM-inkomensvergunning-leads.

Lokaal:  python3 -m uvicorn app:app --port 8642   (SQLite)
Vercel:  serverless via api/index.py              (Postgres via DATABASE_URL; toegangscode via APP_ACCESS_CODE)
Import:  twee bestanden per maand — R0443 xlsx (nieuwe vergunningen) + register-CSV (naam/handelsnaam/plaats).
"""
import csv
import io
import json
import os
import re
import unicodedata
from datetime import datetime, date
from pathlib import Path

import openpyxl
import requests
from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response
from pydantic import BaseModel

BASE = Path(__file__).parent

# Lokale .env inlezen (op Vercel komen deze waarden uit de projectinstellingen)
_envfile = BASE / ".env"
if _envfile.exists():
    for _line in _envfile.read_text().splitlines():
        if "=" in _line and not _line.strip().startswith("#"):
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

from database import DB, init_db, PERSISTENT  # noqa: E402

STATUSSEN = ["Nieuw", "Opnieuw binnen", "Geclaimd", "Benaderd", "In gesprek", "Aanstelling",
             "Bestaande relatie", "Afgewezen", "Geen interesse"]
LOPEND = ("Geclaimd", "Benaderd", "In gesprek")
ACCESS_CODE = os.environ.get("APP_ACCESS_CODE")
SERPER_KEY = os.environ.get("SERPER_API_KEY")
CRON_SECRET = os.environ.get("CRON_SECRET")

from mail import stuur_mail, mail_actief  # noqa: E402
from backup import maak_backup, push_naar_github, GEHEIME_INSTELLINGEN  # noqa: E402

DEMO_MODE = bool(os.environ.get("DEMO_MODE"))

app = FastAPI(title="Leadgenerator Felison")

init_db()
if DEMO_MODE:
    from demo_seed import seed_indien_leeg
    seed_indien_leeg()


@app.middleware("http")
async def toegangscode_gate(request: Request, call_next):
    # De cron-route heeft zijn eigen beveiliging (CRON_SECRET) — Vercel Cron
    # kent de teamtoegangscode niet.
    if request.url.path.startswith("/api/cron/"):
        return await call_next(request)
    if ACCESS_CODE and request.url.path.startswith("/api"):
        # Vergelijking is bewust hoofdletter-ongevoelig: telefoons kapitaliseren
        # automatisch en de code wordt mondeling/via appjes doorgegeven.
        gegeven = (request.headers.get("x-toegangscode") or "").strip().lower()
        if gegeven != ACCESS_CODE.strip().lower():
            return JSONResponse({"detail": "Toegangscode vereist"}, status_code=401)
    return await call_next(request)


# ---------- helpers ----------

def norm_naam(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


def lees_instelling(con, sleutel):
    r = con.execute("SELECT waarde FROM instellingen WHERE sleutel=?", (sleutel,)).fetchone()
    return (r["waarde"] or "").strip() if r and r["waarde"] else None


RELATIE_BRONNEN = ("Felison", "Nedasco")


def match_relaties(con):
    """Markeert nog niet gecheckte leads die op naam matchen met een bestaande relatie.

    Een kantoor kan in meerdere bronlijsten staan; de bronnen worden dan
    samengevoegd ("Felison + Nedasco"). Leads waarvan de check al is afgerond
    ('geen' of 'bevestigd') worden nooit opnieuw gemarkeerd.
    """
    relaties = {}
    for r in con.execute("SELECT naam, naam_norm, bron FROM relaties"):
        relaties.setdefault(r["naam_norm"], {"naam": r["naam"], "bronnen": []})["bronnen"].append(r["bron"] or "?")
    if not relaties:
        return 0
    nieuw = 0
    for l in list(con.execute("SELECT id, naam FROM leads WHERE relatie_match IS NULL")):
        rel = relaties.get(norm_naam(l["naam"]))
        if rel:
            con.execute("UPDATE leads SET relatie_match='mogelijk', relatie_naam=?, relatie_bron=? WHERE id=?",
                        (rel["naam"], " + ".join(sorted(set(rel["bronnen"]))), l["id"]))
            nieuw += 1
    return nieuw


def clean_plaats(p):
    """'Oosterblokker, Gemeente Drechterland' -> 'Oosterblokker'; 'Gemeente Rotterdam' -> 'Rotterdam'."""
    if not p:
        return p
    p = p.split(",")[0].strip()
    p = re.split(r"\s+in\s+de\s+gemeente\b", p, flags=re.IGNORECASE)[0].strip()
    if p.lower().startswith("gemeente "):
        p = p[len("gemeente "):]
    return p.strip() or None


def parse_adres(adres: str):
    """'De Weidenweg 9 7961LN Ruinerwold NL' -> (straat, postcode, plaats)"""
    if not adres:
        return None, None, None
    m = re.search(r"(\d{4}\s?[A-Z]{2})\s+(.+?)(?:\s+NL)?$", adres.strip())
    if m:
        straat = adres[: m.start()].strip()
        return straat or None, m.group(1).replace(" ", ""), m.group(2).strip()
    return adres.strip(), None, None


def geocode(query: str):
    """PDOK Locatieserver (gratis, geen sleutel). Geeft (provincie, lat, lon) of (None,)*3."""
    try:
        r = requests.get(
            "https://api.pdok.nl/bzk/locatieserver/search/v3_1/free",
            params={"q": query, "rows": 1, "fl": "provincienaam,centroide_ll"},
            timeout=6,
        )
        docs = r.json().get("response", {}).get("docs", [])
        if docs:
            prov = docs[0].get("provincienaam")
            m = re.search(r"POINT\(([\d.]+) ([\d.]+)\)", docs[0].get("centroide_ll", ""))
            if m:
                return prov, float(m.group(2)), float(m.group(1))
            return prov, None, None
    except Exception:
        pass
    return None, None, None


def bereken_score(lead: dict):
    """Kwaliteitsscore 0-100 op basis van AFM-data. Geeft (score, klasse, uitleg)."""
    score, uitleg = 40, []
    bd = lead.get("begindatum_dienst")
    if bd:
        dagen = (date.today() - date.fromisoformat(bd)).days
        if dagen <= 45:
            score += 25; uitleg.append("verse vergunning (+25)")
        elif dagen <= 120:
            score += 15; uitleg.append("recente vergunning (+15)")
        else:
            uitleg.append("oudere vergunning (+0)")
    if lead.get("dienst") == "Adviseren / Bemiddelen":
        score += 10; uitleg.append("adviseren/bemiddelen (+10)")
    else:
        score -= 10; uitleg.append(f"afwijkende dienst: {lead.get('dienst')} (-10)")
    if lead.get("beperkingen"):
        score -= 30; uitleg.append("beperkte vergunning, bijv. alleen betalingsbeschermers (-30)")
    rv = (lead.get("rechtsvorm") or "").lower()
    if "besloten" in rv or "naamloze" in rv or "b.v." in (lead.get("naam") or "").lower():
        score += 10; uitleg.append("B.V./N.V. (+10)")
    if lead.get("postcode"):
        score += 10; uitleg.append("volledig adres bekend (+10)")
    elif lead.get("plaats"):
        score += 5; uitleg.append("vestigingsplaats bekend (+5)")
    score = max(0, min(100, score))
    klasse = "A" if score >= 70 else "B" if score >= 50 else "C"
    return score, klasse, "; ".join(uitleg)


# ---------- import ----------

@app.post("/api/import")
async def importeer(xlsx: UploadFile = File(...), register_csv: UploadFile = File(...)):
    # 1. R0443 xlsx inlezen
    try:
        wb = openpyxl.load_workbook(io.BytesIO(await xlsx.read()), read_only=True)
    except Exception:
        raise HTTPException(400, "Kon het xlsx-bestand niet lezen — is dit het R0443-rapport?")
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {str(h).strip(): n for n, h in enumerate(rows[0]) if h}
    verplicht = ["Vergunningnummer", "Naam", "Begindatum dienst"]
    if any(k not in hdr for k in verplicht):
        raise HTTPException(400, f"Xlsx mist kolommen; verwacht o.a. {verplicht}")

    leads = []
    for r in rows[1:]:
        if not r[hdr["Vergunningnummer"]]:
            continue  # lege opmaakrijen in het AFM-rapport
        def g(k):
            v = r[hdr[k]] if k in hdr else None
            return None if v in (None, "-", "") else v
        bd = g("Begindatum dienst")
        leads.append({
            "vergunningnummer": str(g("Vergunningnummer")),
            "naam": str(g("Naam")),
            "rechtsvorm": g("Rechtsvorm"),
            "kvk": str(g("KvK-nummer")) if g("KvK-nummer") else None,
            "adres_raw": g("Adres"),
            "dienst": g("Dienst"),
            "beperkingen": g("Beperkingen vergunning"),
            "begindatum_vergunning": g("Begindatum vergunning").date().isoformat() if isinstance(g("Begindatum vergunning"), datetime) else None,
            "begindatum_dienst": bd.date().isoformat() if isinstance(bd, datetime) else None,
        })

    # 2. register-CSV inlezen (naam;handelsnaam;plaats — cp1252, ;-gescheiden)
    raw = await register_csv.read()
    try:
        tekst = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        tekst = raw.decode("cp1252", errors="replace")
    register = {}
    rd = csv.reader(io.StringIO(tekst), delimiter=";")
    next(rd, None)
    for row in rd:
        if len(row) >= 3:
            register.setdefault(norm_naam(row[0]), row)

    con = DB()
    bestaand = {r["vergunningnummer"] for r in con.execute("SELECT vergunningnummer FROM leads")}

    import_id = con.insert_id("INSERT INTO imports(bestanden, nieuw, dubbel, gematcht) VALUES(?,0,0,0)",
                              (f"{xlsx.filename} + {register_csv.filename}",))

    nieuw = dubbel = gematcht = heropend = 0
    opnieuw_binnen = []
    for l in leads:
        # Elke binnenkomst wordt vastgelegd, ook van al bekende leads — zo is
        # later te zien hoe vaak en wanneer een kantoor in de AFM-lijsten zat.
        con.execute("INSERT INTO lead_historie(vergunningnummer, naam, import_id) VALUES(?,?,?)",
                    (l["vergunningnummer"], l["naam"], import_id))
        if l["vergunningnummer"] in bestaand:
            dubbel += 1
            opnieuw_binnen.append(l["naam"])
            # Eerder afgesloten leads automatisch heropenen: opnieuw in de AFM-lijst
            # verschijnen betekent meestal een vergunningswijziging — nieuw contactmoment.
            oud = con.execute("SELECT id, status FROM leads WHERE vergunningnummer=?",
                              (l["vergunningnummer"],)).fetchone()
            if oud and oud["status"] in ("Afgewezen", "Geen interesse"):
                con.execute("UPDATE leads SET status='Opnieuw binnen' WHERE id=?", (oud["id"],))
                con.execute("INSERT INTO status_log(lead_id, status, am, notitie) VALUES(?,?,?,?)",
                            (oud["id"], "Opnieuw binnen", None,
                             "Automatisch heropend: kantoor kwam opnieuw voor in het AFM-bestand"))
                heropend += 1
            continue
        straat, postcode, plaats = parse_adres(l["adres_raw"])
        handelsnamen = None
        reg = register.get(norm_naam(l["naam"]))
        if reg:
            gematcht += 1
            handelsnamen = reg[1] or None
            if not plaats and reg[2]:
                plaats = reg[2].strip().title()
        plaats = clean_plaats(plaats)
        # geocoderen: postcode > plaats
        provincie = lat = lon = None
        q = f"{postcode} {plaats}" if postcode else plaats
        if q:
            provincie, lat, lon = geocode(q)
        l.update({"adres": straat, "postcode": postcode, "plaats": plaats})
        score, klasse, uitleg = bereken_score(l)
        con.execute("""INSERT INTO leads(vergunningnummer, naam, handelsnamen, rechtsvorm, kvk,
            adres, postcode, plaats, provincie, lat, lon, dienst, beperkingen,
            begindatum_vergunning, begindatum_dienst, score, score_basis, klasse, score_uitleg, import_id)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (l["vergunningnummer"], l["naam"], handelsnamen, l["rechtsvorm"], l["kvk"],
             straat, postcode, plaats, provincie, lat, lon, l["dienst"], l["beperkingen"],
             l["begindatum_vergunning"], l["begindatum_dienst"], score, score, klasse, uitleg, import_id))
        nieuw += 1

    con.execute("UPDATE imports SET nieuw=?, dubbel=?, gematcht=? WHERE id=?",
                (nieuw, dubbel, gematcht, import_id))
    relatie_matches = match_relaties(con)
    con.commit(); con.close()
    return {"nieuw": nieuw, "dubbel_overgeslagen": dubbel, "gematcht_met_register": gematcht,
            "totaal_in_xlsx": len(leads), "opnieuw_binnengekomen": opnieuw_binnen[:15],
            "heropend": heropend, "mogelijke_relaties": relatie_matches}


@app.post("/api/onderhoud/plaatsnamen")
def schoon_plaatsnamen():
    """Eenmalig/idempotent: gemeente-toevoegingen uit bestaande plaatsnamen halen."""
    con = DB()
    aangepast = 0
    for r in list(con.execute("SELECT id, plaats FROM leads WHERE plaats IS NOT NULL")):
        schoon = clean_plaats(r["plaats"])
        if schoon != r["plaats"]:
            con.execute("UPDATE leads SET plaats=? WHERE id=?", (schoon, r["id"]))
            aangepast += 1
    con.commit(); con.close()
    return {"aangepast": aangepast}


# ---------- leads & opvolging ----------

@app.get("/api/leads")
def get_leads(status: str = None, klasse: str = None, provincie: str = None, am: str = None):
    q, p = ("SELECT l.*, h.keren_binnen, h.laatst_binnen FROM leads l "
            "LEFT JOIN (SELECT vergunningnummer, COUNT(*) keren_binnen, MAX(ts) laatst_binnen "
            "FROM lead_historie GROUP BY vergunningnummer) h "
            "ON h.vergunningnummer = l.vergunningnummer WHERE 1=1"), []
    for veld, waarde in [("status", status), ("klasse", klasse), ("provincie", provincie), ("am", am)]:
        if waarde:
            q += f" AND l.{veld}=?"; p.append(waarde)
    q += " ORDER BY l.score DESC, l.naam"
    con = DB()
    out = list(con.execute(q, p))
    con.close()
    return out


class ClaimBody(BaseModel):
    am: str


@app.post("/api/leads/{lead_id}/claim")
def claim(lead_id: int, body: ClaimBody):
    con = DB()
    lead = con.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
    if not lead:
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    if lead["am"] and lead["am"] != body.am:
        con.close()
        raise HTTPException(409, f"Al geclaimd door {lead['am']}")
    eerste_claim = lead["am"] is None
    con.execute("UPDATE leads SET am=?, status=CASE WHEN status='Nieuw' THEN 'Geclaimd' ELSE status END WHERE id=?",
                (body.am, lead_id))
    con.execute("INSERT INTO status_log(lead_id, status, am, notitie) VALUES(?,?,?,?)",
                (lead_id, "Geclaimd", body.am, None))
    marketing = lees_instelling(con, "marketing_email")
    con.commit(); con.close()

    mail_resultaat = None
    if eerste_claim:
        # Marketing (Nicky) direct informeren: haar vervolgstap is het presentje.
        # Een mailfout mag de claim nooit blokkeren.
        try:
            adresregel = ", ".join(x for x in [lead["adres"], lead["postcode"], lead["plaats"]] if x) or "adres onbekend"
            mail_resultaat = stuur_mail(
                marketing,
                f"Nieuwe lead geclaimd: {lead['naam']}",
                "Nieuwe lead geclaimd — presentje versturen",
                [f"<b>{lead['naam']}</b> ({adresregel}) is zojuist geclaimd door <b>{body.am}</b>.",
                 f"Contactpersoon: {lead['contactpersoon'] or 'onbekend'} · {lead['telefoon'] or 'geen telefoon'} · {lead['email'] or 'geen e-mail'}",
                 "Registreer het presentje in het leaddetail van de app, of gebruik de werklijst-export onder Instellingen."],
                "Open het leaddetail")
        except Exception:
            pass
    return {"ok": True, "mail": mail_resultaat}


@app.post("/api/leads/{lead_id}/vrijgeven")
def vrijgeven(lead_id: int):
    con = DB()
    if not con.execute("SELECT 1 FROM leads WHERE id=?", (lead_id,)).fetchone():
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    con.execute("UPDATE leads SET am=NULL, status='Nieuw' WHERE id=?", (lead_id,))
    con.execute("INSERT INTO status_log(lead_id, status, am, notitie) VALUES(?,?,?,?)",
                (lead_id, "Nieuw", None, "Lead vrijgegeven"))
    con.commit(); con.close()
    return {"ok": True}


class StatusBody(BaseModel):
    status: str
    notitie: str = None
    am: str = None


@app.post("/api/leads/{lead_id}/status")
def zet_status(lead_id: int, body: StatusBody):
    if body.status not in STATUSSEN:
        raise HTTPException(400, f"Ongeldige status; kies uit {STATUSSEN}")
    con = DB()
    if not con.execute("SELECT 1 FROM leads WHERE id=?", (lead_id,)).fetchone():
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    con.execute("UPDATE leads SET status=? WHERE id=?", (body.status, lead_id))
    con.execute("INSERT INTO status_log(lead_id, status, am, notitie) VALUES(?,?,?,?)",
                (lead_id, body.status, body.am, body.notitie))
    con.commit(); con.close()
    return {"ok": True}


class WebsiteBody(BaseModel):
    website: str


@app.post("/api/leads/{lead_id}/website")
def zet_website(lead_id: int, body: WebsiteBody):
    con = DB()
    con.execute("UPDATE leads SET website=? WHERE id=?", (body.website, lead_id))
    con.commit(); con.close()
    return {"ok": True}


# Zakelijke inkomensthema's: dit zijn de diensten waarop we een samenwerkingskans beoordelen.
THEMAS = [
    ("verzuimverzekering", 3, [r"verzuimverzekering", r"ziekteverzuim", r"\bverzuim\b"]),
    ("collectieve inkomensverzekeringen", 3, [r"collectieve?\s+inkomens", r"collectieve?\s+verzuim",
                                              r"collectieve?\s+wia", r"collectieve?\s+arbeidsongeschiktheid"]),
    ("personeelsverzekeringen", 3, [r"personeelsverzekering", r"werknemersverzekering",
                                    r"personeel\s+verzeker", r"werkgeversverzekering"]),
    ("AOV/WIA/WGA", 2, [r"\baov\b", r"\bwia\b", r"\bwga\b", r"arbeidsongeschiktheid"]),
]
LINK_HINTS = ["verzuim", "inkomen", "collectief", "personeel", "zakelijk", "werkgever",
              "aov", "wia", "dienst", "verzeker"]
TEAM_WOORDEN = ["team", "medewerkers", "over ons", "adviseurs", "wie zijn wij", "contact"]
UA = {"User-Agent": "Mozilla/5.0 (compatible; LeadCheck/1.0)"}


def scan_site(url: str):
    """Haalt homepage + max 3 relevante subpagina's op. Geeft (punten 0-15, uitleg-string)."""
    from urllib.parse import urljoin, urlparse
    try:
        r = requests.get(url, timeout=8, headers=UA, allow_redirects=True)
        r.raise_for_status()
    except Exception:
        return 0, "website niet bereikbaar (+0)"

    paginas = [r.text.lower()]
    basis_host = urlparse(r.url).netloc
    links = re.findall(r'href=["\']([^"\'#?]+)', r.text, flags=re.IGNORECASE)
    relevant, gezien = [], set()
    for link in links:
        vol = urljoin(r.url, link)
        if urlparse(vol).netloc != basis_host or vol in gezien:
            continue
        if any(h in vol.lower() for h in LINK_HINTS):
            gezien.add(vol)
            relevant.append(vol)
        if len(relevant) >= 3:
            break
    for sub in relevant:
        try:
            sr = requests.get(sub, timeout=6, headers=UA)
            if sr.ok:
                paginas.append(sr.text.lower())
        except Exception:
            pass

    tekst = " ".join(paginas)
    punten, uitleg, gevonden = 3, [f"bereikbaar, {len(paginas)} pagina('s) bekeken (+3)"], []
    for naam, waarde, patronen in THEMAS:
        if any(re.search(p, tekst) for p in patronen):
            punten += waarde
            gevonden.append(f"{naam} (+{waarde})")
    if gevonden:
        uitleg.append("actief op: " + ", ".join(gevonden))
    else:
        uitleg.append("geen zakelijke inkomensdiensten gevonden (+0)")
    if any(w in tekst for w in TEAM_WOORDEN):
        punten += 1
        uitleg.append("team-/contactpagina (+1)")
    return min(15, punten), "; ".join(uitleg)


# Gidsen, registers en socials zijn nooit de eigen site van het kantoor.
ZOEK_BLOCKLIST = ["linkedin.", "facebook.", "instagram.", "youtube.", "twitter.", "x.com",
                  "kvk.nl", "afm.nl", "telefoonboek", "drimble", "openkvk", "google.",
                  "indeed", "glassdoor", "cylex", "oozo.nl", "bedrijvenpagina", "wikipedia",
                  "werkzoeken", "trustoo", "advieskeuze", "independer", "marktplaats",
                  "companyinfo", "company.info", "allebedrijven", "bedrijvenregister",
                  "detelefoongids", "goudengids", "graydon", "creditsafe", "opencompanies",
                  "onderneming.", "bedrijfspagina", "firmania", "infobel"]


def zoek_website(naam: str, plaats: str):
    """Zoekt de eigen website van een kantoor via Serper. Geeft URL of None."""
    if not SERPER_KEY:
        return None
    from urllib.parse import urlparse
    try:
        r = requests.post("https://google.serper.dev/search",
                          json={"q": f"{naam} {plaats or ''}".strip(), "gl": "nl", "hl": "nl", "num": 5},
                          headers={"X-API-KEY": SERPER_KEY, "Content-Type": "application/json"},
                          timeout=10)
        for res in r.json().get("organic", []):
            link = res.get("link", "")
            host = urlparse(link).netloc.lower()
            if host and not any(b in host for b in ZOEK_BLOCKLIST):
                return f"{urlparse(link).scheme}://{host}"
    except Exception:
        pass
    return None


def pas_scan_toe(con, lead, website, punten, uitleg):
    basis = lead["score_basis"] if lead["score_basis"] is not None else lead["score"]
    nieuw = max(0, min(100, basis + punten))
    klasse = "A" if nieuw >= 70 else "B" if nieuw >= 50 else "C"
    con.execute("UPDATE leads SET website=?, website_score=?, website_uitleg=?, score=?, score_basis=?, klasse=? WHERE id=?",
                (website, punten, uitleg, nieuw, basis, klasse, lead["id"]))
    return nieuw, klasse


@app.post("/api/verrijk")
def verrijk(max_leads: int = 5):
    """Zoekt en scant websites voor leads die nog niet verrijkt zijn (batch, herhaald aanroepen tot resterend=0)."""
    if not SERPER_KEY:
        raise HTTPException(400, "Geen SERPER_API_KEY geconfigureerd")
    onverwerkt_sql = ("(website IS NULL AND website_uitleg IS NULL) "
                      "OR (website IS NOT NULL AND website_score IS NULL)")
    con = DB()
    kandidaten = list(con.execute(
        f"SELECT * FROM leads WHERE {onverwerkt_sql} ORDER BY id LIMIT ?", (max_leads,)))
    verwerkt = []
    for lead in kandidaten:
        website = lead["website"] or zoek_website(lead["naam"], lead["plaats"])
        if not website:
            con.execute("UPDATE leads SET website_score=0, website_uitleg='geen website gevonden (+0)' WHERE id=?",
                        (lead["id"],))
            verwerkt.append({"naam": lead["naam"], "website": None, "punten": 0})
            continue
        punten, uitleg = scan_site(website)
        pas_scan_toe(con, lead, website, punten, uitleg)
        verwerkt.append({"naam": lead["naam"], "website": website, "punten": punten})
    con.commit()
    resterend = con.execute(f"SELECT COUNT(*) n FROM leads WHERE {onverwerkt_sql}").fetchone()["n"]
    con.close()
    return {"verwerkt": verwerkt, "resterend": resterend}


@app.post("/api/leads/{lead_id}/scan_website")
def scan_website(lead_id: int):
    """Scant de ingevulde website van de lead en telt het resultaat (0 tot +15) mee in de score."""
    con = DB()
    lead = con.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
    if not lead:
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    if not lead["website"]:
        con.close()
        raise HTTPException(400, "Vul eerst een website in bij deze lead")

    url = lead["website"].strip()
    if not url.startswith("http"):
        url = "https://" + url
    punten, scan_uitleg = scan_site(url)
    uitleg = [scan_uitleg]

    basis = lead["score_basis"] if lead["score_basis"] is not None else lead["score"]
    nieuw = max(0, min(100, basis + punten))
    klasse = "A" if nieuw >= 70 else "B" if nieuw >= 50 else "C"
    con.execute("UPDATE leads SET website_score=?, website_uitleg=?, score=?, score_basis=?, klasse=? WHERE id=?",
                (punten, "; ".join(uitleg), nieuw, basis, klasse, lead_id))
    con.commit(); con.close()
    return {"website_score": punten, "uitleg": "; ".join(uitleg), "nieuwe_score": nieuw, "klasse": klasse}


class GegevensBody(BaseModel):
    telefoon: str = None
    email: str = None
    contactpersoon: str = None


@app.post("/api/leads/{lead_id}/gegevens")
def zet_gegevens(lead_id: int, body: GegevensBody):
    con = DB()
    con.execute("UPDATE leads SET telefoon=?, email=?, contactpersoon=? WHERE id=?",
                (body.telefoon or None, body.email or None, body.contactpersoon or None, lead_id))
    con.commit(); con.close()
    return {"ok": True}


class VervolgBody(BaseModel):
    datum: str = None
    actie: str = None


@app.post("/api/leads/{lead_id}/vervolg")
def zet_vervolg(lead_id: int, body: VervolgBody):
    con = DB()
    con.execute("UPDATE leads SET vervolg_datum=?, vervolg_actie=? WHERE id=?",
                (body.datum or None, body.actie or None, lead_id))
    con.commit(); con.close()
    return {"ok": True}


# Geclaimde leads die nog een presentje van marketing moeten krijgen.
OPENSTAAND_PRESENTJE_SQL = ("am IS NOT NULL AND presentje_datum IS NULL "
                            "AND status NOT IN ('Afgewezen','Geen interesse')")


def norm_datum(v):
    """Excel-cel of tekst naar 'jjjj-mm-dd'; None als het geen datum is."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    m = re.fullmatch(r"(\d{1,2})-(\d{1,2})-(\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


def registreer_presentje(con, lead_id, datum, soort, vervolg_datum=None, vervolg_actie=None):
    """Zet presentje-velden, optioneel de vervolgactie voor de AM, en logt een contactmoment."""
    con.execute("UPDATE leads SET presentje_datum=?, presentje_type=? WHERE id=?",
                (datum, soort, lead_id))
    if vervolg_datum or vervolg_actie:
        con.execute("UPDATE leads SET vervolg_datum=?, vervolg_actie=? WHERE id=?",
                    (vervolg_datum or None, vervolg_actie or None, lead_id))
    con.execute("INSERT INTO contactmomenten(lead_id, type, notitie, am) VALUES(?,?,?,?)",
                (lead_id, "Presentje", f"{soort} verstuurd", "Marketing"))


class PresentjeBody(BaseModel):
    datum: str
    type: str
    vervolg_datum: str = None
    vervolg_actie: str = None


@app.post("/api/leads/{lead_id}/presentje")
def zet_presentje(lead_id: int, body: PresentjeBody):
    datum = norm_datum(body.datum)
    if not datum:
        raise HTTPException(400, "Ongeldige datum; gebruik jjjj-mm-dd")
    con = DB()
    if not con.execute("SELECT 1 FROM leads WHERE id=?", (lead_id,)).fetchone():
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    soorten = {r["naam"] for r in con.execute("SELECT naam FROM presentje_types")}
    if body.type not in soorten:
        con.close()
        raise HTTPException(400, f"Onbekend soort presentje '{body.type}' — beheer de lijst via Instellingen")
    registreer_presentje(con, lead_id, datum, body.type,
                         norm_datum(body.vervolg_datum), body.vervolg_actie)
    con.commit(); con.close()
    return {"ok": True}


PRESENTJE_EXPORT_KOP = ["Vergunningnr", "Naam", "Plaats", "AM", "Adres", "Postcode", "Contactpersoon",
                        "Presentje soort", "Verstuurd op (jjjj-mm-dd)",
                        "Vervolgactie AM", "Vervolgdatum (jjjj-mm-dd)"]


@app.get("/api/presentjes/export")
def presentjes_export():
    """Werklijst voor marketing: geclaimde leads zonder presentje, met invulkolommen."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    con = DB()
    rows = list(con.execute(f"SELECT * FROM leads WHERE {OPENSTAAND_PRESENTJE_SQL} ORDER BY am, naam"))
    con.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Presentjes"
    ws.append(PRESENTJE_EXPORT_KOP)
    for cel in ws[1]:
        cel.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([r["vergunningnummer"], r["naam"], r["plaats"], r["am"], r["adres"],
                   r["postcode"], r["contactpersoon"], None, None, None, None])
    for i, breedte in enumerate([14, 40, 18, 14, 26, 10, 18, 18, 22, 26, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = breedte
    buf = io.BytesIO()
    wb.save(buf)
    naam = f"presentjes-{date.today().isoformat()}.xlsx"
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{naam}"'})


@app.post("/api/presentjes/import")
async def presentjes_import(xlsx: UploadFile = File(...)):
    """Ingevulde presentjes-export terug inlezen: matcht op vergunningnummer,
    verwerkt rijen met soort + verstuurd-op, en meldt fouten per rij terug."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(await xlsx.read()), read_only=True)
    except Exception:
        raise HTTPException(400, "Kon het xlsx-bestand niet lezen — is dit de presentjes-export?")
    rows = list(wb.worksheets[0].iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Leeg bestand")
    # Kolommen op kop-tekst zoeken zodat marketing kolommen mag verplaatsen/verwijderen.
    kop = {str(h or "").strip().lower(): n for n, h in enumerate(rows[0])}
    def kol(*zoek):
        return next((n for k, n in kop.items() if any(z in k for z in zoek)), None)
    k_vergunning, k_soort, k_datum = kol("vergunningnr"), kol("presentje soort"), kol("verstuurd op")
    k_vactie, k_vdatum = kol("vervolgactie"), kol("vervolgdatum")
    if k_vergunning is None or k_soort is None or k_datum is None:
        raise HTTPException(400, "Bestand mist kolommen; verwacht o.a. Vergunningnr, Presentje soort, Verstuurd op")

    con = DB()
    soorten = {r["naam"] for r in con.execute("SELECT naam FROM presentje_types")}
    verwerkt, overgeslagen, fouten = 0, 0, []
    for r in rows[1:]:
        def cel(n):
            v = r[n] if n is not None and n < len(r) else None
            return str(v).strip() if isinstance(v, str) else v
        vergunning = cel(k_vergunning)
        if vergunning is None:
            continue  # lege rij
        vergunning = str(vergunning)
        soort, datum_raw = cel(k_soort), cel(k_datum)
        if not soort and not datum_raw:
            overgeslagen += 1  # nog niet ingevuld door marketing
            continue
        lead = con.execute("SELECT id, naam FROM leads WHERE vergunningnummer=?", (vergunning,)).fetchone()
        if not lead:
            fouten.append(f"{vergunning}: geen lead met dit vergunningnummer")
            continue
        datum = norm_datum(datum_raw)
        if not soort or not datum:
            fouten.append(f"{lead['naam']}: soort én geldige verstuurd-op-datum (jjjj-mm-dd) zijn verplicht")
            continue
        if soort not in soorten:
            fouten.append(f"{lead['naam']}: onbekend soort '{soort}' — voeg het eerst toe via Instellingen")
            continue
        registreer_presentje(con, lead["id"], datum, soort,
                             norm_datum(cel(k_vdatum)), cel(k_vactie) or None)
        verwerkt += 1
    con.commit(); con.close()
    return {"verwerkt": verwerkt, "overgeslagen": overgeslagen, "fouten": fouten}


class PresentjeTypeBody(BaseModel):
    naam: str


@app.get("/api/presentje_types")
def get_presentje_types():
    con = DB()
    out = [r["naam"] for r in con.execute("SELECT naam FROM presentje_types ORDER BY naam")]
    con.close()
    return out


@app.post("/api/presentje_types")
def add_presentje_type(body: PresentjeTypeBody):
    naam = body.naam.strip()
    if not naam:
        raise HTTPException(400, "Naam is verplicht")
    con = DB()
    if con.pg:
        con.execute("INSERT INTO presentje_types(naam) VALUES(?) ON CONFLICT(naam) DO NOTHING", (naam,))
    else:
        con.execute("INSERT OR IGNORE INTO presentje_types(naam) VALUES(?)", (naam,))
    con.commit(); con.close()
    return {"ok": True}


@app.delete("/api/presentje_types/{naam}")
def del_presentje_type(naam: str):
    # Reeds geregistreerde presentjes op leads behouden hun soort (historie).
    con = DB()
    con.execute("DELETE FROM presentje_types WHERE naam=?", (naam,))
    con.commit(); con.close()
    return {"ok": True}


class ContactBody(BaseModel):
    type: str
    notitie: str = None
    am: str = None


@app.post("/api/leads/{lead_id}/contact")
def log_contact(lead_id: int, body: ContactBody):
    con = DB()
    if not con.execute("SELECT 1 FROM leads WHERE id=?", (lead_id,)).fetchone():
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    con.execute("INSERT INTO contactmomenten(lead_id, type, notitie, am) VALUES(?,?,?,?)",
                (lead_id, body.type, body.notitie, body.am))
    con.commit(); con.close()
    return {"ok": True}


@app.get("/api/leads/{lead_id}/activiteit")
def get_activiteit(lead_id: int):
    """Gecombineerde tijdlijn: statuswissels + contactmomenten, nieuwste eerst."""
    con = DB()
    feed = [{"soort": "status", "label": r["status"], "notitie": r["notitie"], "am": r["am"], "ts": str(r["ts"])}
            for r in con.execute("SELECT * FROM status_log WHERE lead_id=?", (lead_id,))]
    feed += [{"soort": "contact", "label": r["type"], "notitie": r["notitie"], "am": r["am"], "ts": str(r["ts"])}
             for r in con.execute("SELECT * FROM contactmomenten WHERE lead_id=?", (lead_id,))]
    con.close()
    return sorted(feed, key=lambda x: x["ts"], reverse=True)


def _dt(v):
    if isinstance(v, datetime):
        return v
    return datetime.fromisoformat(str(v).replace(" ", "T").split(".")[0])


@app.get("/api/export")
def export_excel(status: str = None, klasse: str = None, provincie: str = None, am: str = None):
    """Excel-export van de leadlijst, met dezelfde filters als het dashboard."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    rows = get_leads(status, klasse, provincie, am)
    kolommen = [
        ("Vergunningnr", "vergunningnummer", 14), ("Naam", "naam", 40), ("Plaats", "plaats", 18),
        ("Provincie", "provincie", 14), ("Score", "score", 8), ("Klasse", "klasse", 8),
        ("Status", "status", 15), ("AM", "am", 14), ("Contactpersoon", "contactpersoon", 18),
        ("Telefoon", "telefoon", 14), ("E-mail", "email", 24), ("Website", "website", 28),
        ("Vervolgactie", "vervolg_actie", 18), ("Vervolgdatum", "vervolg_datum", 13),
        ("Presentje", "presentje_type", 16), ("Presentje verstuurd", "presentje_datum", 14),
        ("Vergunning per", "begindatum_dienst", 14), ("Keren binnengekomen", "keren_binnen", 12),
        ("KvK", "kvk", 11), ("Rechtsvorm", "rechtsvorm", 18), ("Adres", "adres", 26),
        ("Postcode", "postcode", 10), ("Beperkingen", "beperkingen", 40),
        ("Score-uitleg", "score_uitleg", 50), ("Websitescan", "website_uitleg", 50),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append([k[0] for k in kolommen])
    for cel in ws[1]:
        cel.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([r.get(veld) for _, veld, _ in kolommen])
    for i, (_, _, breedte) in enumerate(kolommen, 1):
        ws.column_dimensions[get_column_letter(i)].width = breedte
    buf = io.BytesIO()
    wb.save(buf)
    naam = f"leads-{date.today().isoformat()}.xlsx"
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{naam}"'})


@app.get("/api/conversie")
def conversie():
    """Funnel per maandcohort, conversie per scoreklasse en doorlooptijden."""
    con = DB()
    leads = list(con.execute(
        "SELECT id, klasse, status, aangemaakt, am, presentje_datum, presentje_type FROM leads"))
    log = list(con.execute("SELECT lead_id, status, ts FROM status_log"))
    con.close()

    ooit = {}  # lead_id -> {status: eerste ts}
    for r in log:
        ooit.setdefault(r["lead_id"], {}).setdefault(r["status"], r["ts"])

    def bereikt(l, *statussen):
        return l["status"] in statussen or any(s in ooit.get(l["id"], {}) for s in statussen)

    cohorten = {}
    for l in leads:
        maand = str(l["aangemaakt"])[:7]
        c = cohorten.setdefault(maand, {"maand": maand, "leads": 0, "benaderd": 0,
                                        "in_gesprek": 0, "aanstelling": 0, "afgesloten": 0})
        c["leads"] += 1
        if bereikt(l, "Benaderd", "In gesprek", "Aanstelling"):
            c["benaderd"] += 1
        if bereikt(l, "In gesprek", "Aanstelling"):
            c["in_gesprek"] += 1
        if bereikt(l, "Aanstelling"):
            c["aanstelling"] += 1
        if l["status"] in ("Afgewezen", "Geen interesse"):
            c["afgesloten"] += 1

    klassen = {}
    for l in leads:
        k = klassen.setdefault(l["klasse"] or "?", {"klasse": l["klasse"] or "?", "leads": 0,
                                                    "benaderd": 0, "aanstelling": 0})
        k["leads"] += 1
        if bereikt(l, "Benaderd", "In gesprek", "Aanstelling"):
            k["benaderd"] += 1
        if bereikt(l, "Aanstelling"):
            k["aanstelling"] += 1
    for k in klassen.values():
        k["conversie_pct"] = round(100 * k["aanstelling"] / k["leads"], 1) if k["leads"] else 0

    tot_aanstelling, tot_actie = [], []
    for l in leads:
        events = ooit.get(l["id"], {})
        start = _dt(l["aangemaakt"])
        if "Aanstelling" in events:
            tot_aanstelling.append((_dt(events["Aanstelling"]) - start).days)
        eerste = min((_dt(ts) for s, ts in events.items() if s in ("Geclaimd", "Benaderd")), default=None)
        if eerste:
            tot_actie.append(max(0, (eerste - start).days))

    # Presentjes: aantallen, verdeling en effect op conversie (alleen geclaimde
    # leads als vergelijking, want een presentje volgt altijd op een claim).
    met_presentje = [l for l in leads if l["presentje_datum"]]
    per_soort, per_maand = {}, {}
    for l in met_presentje:
        per_soort[l["presentje_type"] or "?"] = per_soort.get(l["presentje_type"] or "?", 0) + 1
        maand = str(l["presentje_datum"])[:7]
        per_maand[maand] = per_maand.get(maand, 0) + 1
    openstaand = sum(1 for l in leads if l["am"] and not l["presentje_datum"]
                     and l["status"] not in ("Afgewezen", "Geen interesse"))

    def effect_rij(label, groep):
        n = len(groep)
        gesprek = sum(1 for l in groep if bereikt(l, "In gesprek", "Aanstelling"))
        aanst = sum(1 for l in groep if bereikt(l, "Aanstelling"))
        return {"label": label, "leads": n, "in_gesprek": gesprek, "aanstelling": aanst,
                "conversie_pct": round(100 * aanst / n, 1) if n else 0}

    geclaimd = [l for l in leads if l["am"]]
    presentjes = {
        "verstuurd": len(met_presentje),
        "openstaand": openstaand,
        "per_soort": dict(sorted(per_soort.items(), key=lambda x: -x[1])),
        "per_maand": [{"maand": m, "aantal": n} for m, n in sorted(per_maand.items())],
        "effect": [effect_rij("Met presentje", [l for l in geclaimd if l["presentje_datum"]]),
                   effect_rij("Zonder presentje", [l for l in geclaimd if not l["presentje_datum"]])],
    }

    return {
        "cohorten": sorted(cohorten.values(), key=lambda c: c["maand"]),
        "klassen": sorted(klassen.values(), key=lambda k: k["klasse"]),
        "presentjes": presentjes,
        "doorlooptijd": {
            "gem_dagen_tot_eerste_actie": round(sum(tot_actie) / len(tot_actie), 1) if tot_actie else None,
            "gem_dagen_tot_aanstelling": round(sum(tot_aanstelling) / len(tot_aanstelling), 1) if tot_aanstelling else None,
            "aantal_aanstellingen": len(tot_aanstelling),
        },
    }


@app.get("/api/leads/{lead_id}/historie")
def get_historie(lead_id: int):
    """Alle keren dat deze lead in een AFM-import zat."""
    con = DB()
    lead = con.execute("SELECT vergunningnummer FROM leads WHERE id=?", (lead_id,)).fetchone()
    if not lead:
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    out = list(con.execute(
        "SELECT h.ts, h.import_id, i.bestanden FROM lead_historie h "
        "LEFT JOIN imports i ON i.id = h.import_id "
        "WHERE h.vergunningnummer=? ORDER BY h.ts", (lead["vergunningnummer"],)))
    con.close()
    return out


@app.get("/api/leads/{lead_id}/log")
def get_log(lead_id: int):
    con = DB()
    out = list(con.execute("SELECT * FROM status_log WHERE lead_id=? ORDER BY ts DESC", (lead_id,)))
    con.close()
    return out


# ---------- AM's, stats, meta ----------

class AmBody(BaseModel):
    naam: str
    kleur: str = "#2563eb"
    email: str = None


@app.get("/api/ams")
def get_ams():
    con = DB()
    out = list(con.execute("SELECT * FROM ams ORDER BY naam"))
    con.close()
    return out


@app.post("/api/ams")
def add_am(body: AmBody):
    con = DB()
    con.upsert_am(body.naam.strip(), body.kleur, (body.email or "").strip() or None)
    con.commit(); con.close()
    return {"ok": True}


@app.delete("/api/ams/{naam}")
def del_am(naam: str):
    con = DB()
    con.execute("DELETE FROM ams WHERE naam=?", (naam,))
    con.execute("UPDATE leads SET am=NULL, status=CASE WHEN status='Geclaimd' THEN 'Nieuw' ELSE status END WHERE am=?",
                (naam,))
    con.commit(); con.close()
    return {"ok": True}


# ---------- instellingen, relaties, feedback, digest ----------

class InstellingenBody(BaseModel):
    marketing_email: str = None
    feedback_email: str = None
    github_token: str = None
    backup_repo: str = None


# Deze waarde sturen we terug in plaats van het echte geheim; komt hij terug bij
# een opslag-actie, dan laten we het bestaande geheim ongemoeid.
GEHEIM_MASKER = "••••••••"


@app.get("/api/instellingen")
def get_instellingen():
    con = DB()
    out = {}
    for r in con.execute("SELECT * FROM instellingen"):
        # Geheimen gaan nooit terug naar de browser; alleen of ze gezet zijn.
        out[r["sleutel"]] = GEHEIM_MASKER if r["sleutel"] in GEHEIME_INSTELLINGEN and r["waarde"] else r["waarde"]
    con.close()
    return out


@app.post("/api/instellingen")
def zet_instellingen(body: InstellingenBody):
    con = DB()
    for sleutel in ("marketing_email", "feedback_email", "github_token", "backup_repo"):
        waarde = getattr(body, sleutel)
        if waarde == GEHEIM_MASKER:
            continue  # onveranderd gelaten in de UI
        if waarde is not None:
            if con.pg:
                con.execute("INSERT INTO instellingen(sleutel, waarde) VALUES(?,?) "
                            "ON CONFLICT(sleutel) DO UPDATE SET waarde=EXCLUDED.waarde",
                            (sleutel, waarde.strip()))
            else:
                con.execute("INSERT OR REPLACE INTO instellingen(sleutel, waarde) VALUES(?,?)",
                            (sleutel, waarde.strip()))
    con.commit(); con.close()
    return {"ok": True}


@app.post("/api/relaties/import")
async def relaties_import(bestand: UploadFile = File(...), bron: str = "Felison"):
    """Vervangt de relatielijst van één bron (Felison of Nedasco) door de aangeleverde xlsx/csv."""
    if bron not in RELATIE_BRONNEN:
        raise HTTPException(400, f"Onbekende bron; kies uit {list(RELATIE_BRONNEN)}")
    raw = await bestand.read()
    namen = []
    if bestand.filename.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        rijen = list(wb.worksheets[0].iter_rows(values_only=True))
    else:
        try:
            tekst = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            tekst = raw.decode("cp1252", errors="replace")
        proef = tekst.splitlines()[0] if tekst.splitlines() else ""
        scheiding = ";" if proef.count(";") >= proef.count(",") else ","
        rijen = list(csv.reader(io.StringIO(tekst), delimiter=scheiding))
    if not rijen:
        raise HTTPException(400, "Leeg bestand")
    kolom = 0
    kop = [str(c or "").strip().lower() for c in rijen[0]]
    start = 0
    for i, k in enumerate(kop):
        if k in ("naam", "bedrijfsnaam", "relatie", "kantoor", "statutaire naam"):
            kolom, start = i, 1
            break
    else:
        # geen herkenbare kop: als de eerste cel geen bedrijfsnaam lijkt (bv. 'naam'), toch overslaan
        if kop and kop[0] in ("naam", "name"):
            start = 1
    con = DB()
    con.execute("DELETE FROM relaties WHERE bron=?", (bron,))
    gezien, aantal = set(), 0
    for r in rijen[start:]:
        if len(r) <= kolom or not r[kolom]:
            continue
        naam = str(r[kolom]).strip()
        nn = norm_naam(naam)
        if not nn or nn in gezien:
            continue
        gezien.add(nn)
        con.execute("INSERT INTO relaties(naam, naam_norm, bron) VALUES(?,?,?)", (naam, nn, bron))
        aantal += 1
    # Eerder afgeronde checks blijven staan; alleen open markeringen opnieuw bepalen.
    con.execute("UPDATE leads SET relatie_match=NULL, relatie_naam=NULL, relatie_bron=NULL WHERE relatie_match='mogelijk'")
    matches = match_relaties(con)
    con.commit(); con.close()
    return {"bron": bron, "relaties": aantal, "mogelijke_matches": matches}


@app.get("/api/relaties/status")
def relaties_status():
    con = DB()
    per_bron = {b: 0 for b in RELATIE_BRONNEN}
    for r in con.execute("SELECT bron, COUNT(*) n FROM relaties GROUP BY bron"):
        per_bron[r["bron"] or "?"] = r["n"]
    open_checks = con.execute("SELECT COUNT(*) n FROM leads WHERE relatie_match='mogelijk'").fetchone()["n"]
    con.close()
    return {"per_bron": per_bron, "relaties": sum(per_bron.values()), "te_controleren": open_checks}


class RelatiecheckBody(BaseModel):
    uitkomst: str  # 'bevestigd' of 'geen'
    am: str = None


@app.post("/api/leads/{lead_id}/relatiecheck")
def relatiecheck(lead_id: int, body: RelatiecheckBody):
    if body.uitkomst not in ("bevestigd", "geen"):
        raise HTTPException(400, "uitkomst moet 'bevestigd' of 'geen' zijn")
    con = DB()
    lead = con.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
    if not lead:
        con.close()
        raise HTTPException(404, "Lead niet gevonden")
    con.execute("UPDATE leads SET relatie_match=? WHERE id=?", (body.uitkomst, lead_id))
    if body.uitkomst == "bevestigd":
        con.execute("UPDATE leads SET status='Bestaande relatie' WHERE id=?", (lead_id,))
        con.execute("INSERT INTO status_log(lead_id, status, am, notitie) VALUES(?,?,?,?)",
                    (lead_id, "Bestaande relatie", body.am,
                     f"Bevestigd als bestaande relatie bij {lead['relatie_bron'] or '?'} (match: {lead['relatie_naam']})"))
    else:
        con.execute("INSERT INTO status_log(lead_id, status, am, notitie) VALUES(?,?,?,?)",
                    (lead_id, lead["status"], body.am,
                     f"Relatiecheck: geen match (lijst noemde '{lead['relatie_naam']}')"))
    con.commit(); con.close()
    return {"ok": True}


class FeedbackBody(BaseModel):
    naam: str = None
    tekst: str
    scherm: str = None


@app.post("/api/feedback")
def geef_feedback(body: FeedbackBody):
    if not body.tekst.strip():
        raise HTTPException(400, "Lege feedback")
    con = DB()
    con.execute("INSERT INTO feedback(naam, tekst, scherm) VALUES(?,?,?)",
                (body.naam or "anoniem", body.tekst.strip(), body.scherm))
    ontvanger = lees_instelling(con, "feedback_email")
    con.commit(); con.close()
    mail_resultaat = stuur_mail(
        ontvanger,
        f"App-feedback van {body.naam or 'anoniem'}",
        "Nieuwe feedback op de leadgenerator",
        [f"<b>Van:</b> {body.naam or 'anoniem'} · <b>scherm:</b> {body.scherm or 'onbekend'}",
         f"<i>“{body.tekst.strip()}”</i>"],
        "Open de app")
    return {"ok": True, "mail": mail_resultaat}


@app.get("/api/feedback")
def lijst_feedback():
    con = DB()
    out = list(con.execute("SELECT * FROM feedback ORDER BY id DESC LIMIT 100"))
    con.close()
    return out


@app.get("/api/backup/download")
def backup_download():
    """Handmatige volledige back-up (alle tabellen) als JSON-bestand."""
    con = DB()
    data = maak_backup(con)
    con.close()
    naam = f"leadgenerator-backup-{date.today().isoformat()}.json"
    return Response(json.dumps(data, ensure_ascii=False, indent=1),
                    media_type="application/json",
                    headers={"Content-Disposition": f'attachment; filename="{naam}"'})


def backup_koppeling(con):
    """Token en repo uit de app-instellingen; omgevingsvariabelen als fallback."""
    token = lees_instelling(con, "github_token") or os.environ.get("GITHUB_TOKEN")
    repo = lees_instelling(con, "backup_repo") or os.environ.get("BACKUP_REPO")
    return token, repo


@app.get("/api/cron/backup")
def cron_backup(request: Request):
    """Dagelijkse volledige back-up naar de GitHub-repo (Vercel Cron)."""
    auth = request.headers.get("authorization", "")
    if not CRON_SECRET or auth != f"Bearer {CRON_SECRET}":
        raise HTTPException(401, "Ongeldige cron-authenticatie")
    con = DB()
    data = maak_backup(con)
    token, repo = backup_koppeling(con)
    con.close()
    return push_naar_github(data, token, repo)


@app.post("/api/backup/nu")
def backup_nu():
    """Handmatig een back-up naar GitHub duwen (knop in Instellingen)."""
    con = DB()
    data = maak_backup(con)
    token, repo = backup_koppeling(con)
    con.close()
    return push_naar_github(data, token, repo)


# Volgorde is belangrijk: kindtabellen eerst leeg, ouders (leads) eerst terug.
HERSTEL_VOLGORDE = ["leads", "ams", "relaties", "instellingen", "presentje_types",
                    "imports", "lead_historie", "status_log", "contactmomenten", "feedback"]


@app.post("/api/backup/restore")
async def backup_restore(bestand: UploadFile = File(...), bevestiging: str = ""):
    """Zet een back-upbestand volledig terug. Overschrijft ALLE huidige data."""
    if bevestiging != "HERSTEL":
        raise HTTPException(400, "Herstellen overschrijft alle data. Stuur bevestiging='HERSTEL' mee.")
    try:
        data = json.loads(await bestand.read())
    except Exception:
        raise HTTPException(400, "Geen geldig back-upbestand (JSON verwacht)")
    if data.get("versie") != 1 or "tabellen" not in data:
        raise HTTPException(400, "Onbekend back-upformaat")

    con = DB()
    # De back-up bevat bewust geen geheimen (zie backup.py). Zonder deze stap zou
    # een herstel het GitHub-token wissen en zou de back-up stilletjes stoppen.
    geheimen = [(s, lees_instelling(con, s)) for s in GEHEIME_INSTELLINGEN]
    for tabel in reversed(HERSTEL_VOLGORDE):
        con.execute(f"DELETE FROM {tabel}")
    hersteld = {}
    for tabel in HERSTEL_VOLGORDE:
        rijen = data["tabellen"].get(tabel, [])
        for r in rijen:
            kolommen = list(r.keys())
            plaatsen = ",".join("?" for _ in kolommen)
            con.execute(f"INSERT INTO {tabel} ({','.join(kolommen)}) VALUES({plaatsen})",
                        [r[k] for k in kolommen])
        hersteld[tabel] = len(rijen)
        # Postgres telt zijn id-teller niet vanzelf door bij expliciete id's:
        # zonder dit botst de eerste nieuwe rij op een bestaand id.
        if con.pg and rijen and "id" in rijen[0]:
            con.execute(f"SELECT setval(pg_get_serial_sequence('{tabel}','id'), "
                        f"(SELECT COALESCE(MAX(id),1) FROM {tabel}))")
    for sleutel, waarde in geheimen:
        if waarde:
            con.execute("INSERT INTO instellingen(sleutel, waarde) VALUES(?,?)", (sleutel, waarde))
    con.commit(); con.close()
    return {"hersteld": hersteld, "uit_backup_van": data.get("gemaakt_op")}


@app.get("/api/cron/digest")
def cron_digest(request: Request, test: int = 0):
    """Dagelijkse ochtendmail per AM. Aangeroepen door Vercel Cron (Bearer CRON_SECRET)."""
    auth = request.headers.get("authorization", "")
    if not CRON_SECRET or auth != f"Bearer {CRON_SECRET}":
        raise HTTPException(401, "Ongeldige cron-authenticatie")
    if not test and date.today().weekday() >= 5:  # za/zo overslaan
        return {"overgeslagen": "weekend"}
    vandaag = date.today().isoformat()
    con = DB()
    ams = list(con.execute("SELECT * FROM ams WHERE email IS NOT NULL AND email != ''"))
    resultaat = []
    for am in ams:
        leads = list(con.execute("SELECT * FROM leads WHERE am=?", (am["naam"],)))
        open_lead = lambda l: l["status"] in LOPEND or l["status"] == "Opnieuw binnen"
        achterstallig = [l for l in leads if open_lead(l) and l["vervolg_datum"] and l["vervolg_datum"] < vandaag]
        vandaag_gepland = [l for l in leads if open_lead(l) and l["vervolg_datum"] == vandaag]
        opnieuw = [l for l in leads if l["status"] == "Opnieuw binnen"]
        checks = [l for l in leads if l["relatie_match"] == "mogelijk"]
        if not (achterstallig or vandaag_gepland or opnieuw or checks):
            resultaat.append({"am": am["naam"], "verzonden": False, "reden": "niets te doen"})
            continue
        regels = []
        if achterstallig:
            regels.append("<b>Achterstallig:</b> " + " · ".join(
                f"{l['naam']} ({l['vervolg_actie'] or 'vervolgactie'}, {l['vervolg_datum']})" for l in achterstallig))
        if vandaag_gepland:
            regels.append("<b>Vandaag gepland:</b> " + " · ".join(
                f"{l['naam']} ({l['vervolg_actie'] or 'vervolgactie'})" for l in vandaag_gepland))
        if opnieuw:
            regels.append("<b>Opnieuw binnengekomen:</b> " + " · ".join(l["naam"] for l in opnieuw))
        if checks:
            regels.append("<b>Relatiecheck nodig:</b> " + " · ".join(l["naam"] for l in checks))
        totaal = len(achterstallig) + len(vandaag_gepland) + len(opnieuw) + len(checks)
        r = stuur_mail(am["email"], f"Leadgenerator: {totaal} actie(s) voor vandaag",
                       f"Goedemorgen {am['naam']}, dit staat er voor je klaar", regels)
        resultaat.append({"am": am["naam"], **r})
    con.close()
    return {"datum": vandaag, "resultaat": resultaat}


@app.get("/api/stats")
def stats():
    con = DB()
    out = {
        "per_status": {r["status"]: r["n"] for r in con.execute("SELECT status, COUNT(*) n FROM leads GROUP BY status")},
        "per_klasse": {r["klasse"]: r["n"] for r in con.execute("SELECT klasse, COUNT(*) n FROM leads GROUP BY klasse")},
        "per_am": {r["am"] or "—": r["n"] for r in con.execute("SELECT am, COUNT(*) n FROM leads GROUP BY am")},
        "per_provincie": {r["provincie"] or "?": r["n"] for r in con.execute("SELECT provincie, COUNT(*) n FROM leads GROUP BY provincie")},
        "aanstellingen_per_am": {r["am"] or "—": r["n"] for r in con.execute(
            "SELECT am, COUNT(*) n FROM leads WHERE status='Aanstelling' GROUP BY am")},
        "imports": list(con.execute("SELECT * FROM imports ORDER BY id DESC LIMIT 12")),
    }
    con.close()
    return out


def backup_ingesteld():
    con = DB()
    token, repo = backup_koppeling(con)
    con.close()
    return bool(token and repo)


@app.get("/api/meta")
def meta():
    return {"statussen": STATUSSEN, "persistent": PERSISTENT, "beveiligd": bool(ACCESS_CODE),
            "serper": bool(SERPER_KEY), "demo": DEMO_MODE, "mail": mail_actief(),
            "backup": backup_ingesteld()}


@app.get("/")
def index():
    # Lokaal (uvicorn) serveren we het bestand direct; op Vercel is public/
    # niet altijd in de functiebundel aanwezig — dan sturen we door naar
    # /index.html, dat door het CDN uit public/ wordt geserveerd.
    pad = BASE / "public" / "index.html"
    if pad.exists():
        return FileResponse(pad, headers={"Cache-Control": "no-cache"})
    return RedirectResponse("/index.html", status_code=307)
